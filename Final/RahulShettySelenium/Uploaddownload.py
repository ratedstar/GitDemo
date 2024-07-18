import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


filepath = "C:\\Users\\HP\\Downloads\\download.xlsx"

def update_excel_data(filepath,fruit_name,new_value,column_name):
    book = openpyxl.load_workbook(filepath)
    sheet = book.active
    get_row_column = {}

    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i,column=j).value == fruit_name:
                get_row_column["row"] = i
    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == column_name:
            get_row_column["col"] = i
    sheet.cell(row=get_row_column["row"],column=get_row_column["col"]).value = new_value
    book.save(filepath)


baseUrl = "https://rahulshettyacademy.com/upload-download-test/"
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--ignore-certificate-errors")
driver = webdriver.Chrome(ChromeDriverManager().install(),options=chrome_options)
driver.get(baseUrl)
driver.implicitly_wait(5)
time.sleep(2)
driver.find_element(By.ID,"downloadButton").click()
fruit_name = "Apple"
new_value = 999
column_name = "price"
update_excel_data(filepath,fruit_name,new_value,column_name)
#send file
driver.find_element(By.ID,"fileinput").send_keys("C:\\Users\\HP\\Downloads\\download.xlsx")

wait = WebDriverWait(driver, 5)
alert_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(alert_locator))
alert_message = driver.find_element(By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)").text

price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']//parent::div//parent::div//div[@id='cell-"+price_column+"-undefined']").text
print("The actual price is"+actual_price)
print(type(actual_price))
b = 10
print("{}{}".format("The value of b: ",b))
print("{} {} {}".format("The value of b: ", 10,20,30))
print(f"This is the value of b: {b}")




