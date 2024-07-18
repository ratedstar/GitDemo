import openpyxl

book = openpyxl.load_workbook("C:\\Users\\HP\\Downloads\\PythonSelenium\\pythondemo.xlsx")
sheet = book.active
cell = sheet.cell(row=2, column=2)
print(cell.value)

book = openpyxl.load_workbook("C:\\Users\\HP\\Downloads\\PythonSelenium\\pythondemo.xlsx")
sheet = book.active
cell = sheet.cell(row=2, column=2)
print(cell.value)

#sheet.cell(row=2,column=2).value = "Kaamesh"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet["B4"].value)

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value, end=" ")
print()
test_data = {}
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Testcase2":
        for j in range(2,sheet.max_column+1):
            test_data[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
print([test_data])
test_data = []





