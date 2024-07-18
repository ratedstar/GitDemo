import openpyxl

book = openpyxl.load_workbook("C:\\Users\\HP\\Downloads\\PythonSelenium\\pythondemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=3,column=3).value = "Kaamesh"
print(sheet.cell(row=3,column=3).value)
print(sheet.max_row)
print(sheet.max_column)

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value)
# if we what a particular column data
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Testcase2":
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)

dict = {}
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column + 1):
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(dict)

book = openpyxl.load_workbook("C:\\Users\\HP\\Downloads\\PythonSelenium\\pythondemo.xlsx")
sheet = book.active
cell = sheet.cell(row=3,column=2)
print(cell.value)
sheet.cell(row=3,column=3).value = "Kaamesh"
print(sheet.cell(row=3,column=3).value)
print(sheet.max_row)
print(sheet.max_column)
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value =="Testcase2":
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)
dict={}
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Testcase2":
        for j in range(2,sheet.max_column+1):
            dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
print(dict)

